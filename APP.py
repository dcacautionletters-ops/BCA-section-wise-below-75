import streamlit as st
import pandas as pd
import io
import time
import plotly.express as px
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- 1. UI CONFIGURATION & ADVANCED GLASS-MORPHISM CSS ---
st.set_page_config(page_title="VMS Reporting System", layout="wide")
MASTER_PASSWORD = "VMS@123"

st.markdown("""
    <style>
    .stApp {
        background: linear-gradient(rgba(15, 23, 42, 0.92), rgba(15, 23, 42, 0.92)), 
                    url("https://images.unsplash.com/photo-1451187580459-43490279c0fa?ixlib=rb-1.2.1&auto=format&fit=crop&w=1950&q=80");
        background-size: cover; background-attachment: fixed;
    }
    .welcome-note { 
        background: linear-gradient(to right, #00d2ff, #92fe9d); 
        -webkit-background-clip: text; -webkit-text-fill-color: transparent; 
        font-size: 48px !important; font-weight: 700; text-align: center; margin: 40px 0 10px 0;
    }
    .glass-metric {
        background: rgba(255, 255, 255, 0.03);
        backdrop-filter: blur(15px);
        -webkit-backdrop-filter: blur(15px);
        border: 1px solid rgba(255, 255, 255, 0.1);
        border-radius: 15px;
        padding: 25px;
        margin: 10px 0;
        transition: transform 0.3s ease, background 0.3s ease;
        text-align: center;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
    }
    .glass-metric:hover {
        background: rgba(255, 255, 255, 0.08);
        transform: translateY(-5px);
        border: 1px solid rgba(146, 254, 157, 0.4);
    }
    .metric-title { color: #94a3b8; font-size: 14px; text-transform: uppercase; letter-spacing: 1px; }
    .metric-value { font-size: 42px; font-weight: 800; color: #92fe9d; text-shadow: 0 0 15px rgba(146, 254, 157, 0.3); }
    
    .portal-status {
        font-family: 'Consolas', monospace; color: #92fe9d; font-size: 13px; padding: 15px;
        background: rgba(0, 0, 0, 0.3); border-radius: 4px; border-left: 4px solid #00d2ff; margin-bottom: 15px;
    }
    .footer { position: fixed; bottom: 10px; width: 100%; text-align: center; color: #64748b; font-size: 11px; }
    </style>
    """, unsafe_allow_html=True)

# --- 2. SESSION STATE & AUTHENTICATION ---
if 'authenticated' not in st.session_state: st.session_state.authenticated = False
if 'show_logout_msg' not in st.session_state: st.session_state.show_logout_msg = False
if 'booted' not in st.session_state: st.session_state.booted = False

if st.session_state.show_logout_msg:
    st.markdown('<div style="text-align:center; color:#92fe9d; padding:100px;"><h2>Neural Link Severed. Have a wonderful day! 🌟</h2></div>', unsafe_allow_html=True)
    time.sleep(2.0); st.session_state.show_logout_msg = False; st.session_state.booted = False; st.rerun()

if not st.session_state.authenticated:
    st.markdown('<p class="welcome-note">VMS Reporting System</p>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1.4, 1])
    with col2:
        if not st.session_state.booted:
            st.markdown('<div class="portal-status">> Initiating Secure Link...<br>> Encryption Layers Active.</div>', unsafe_allow_html=True)
            time.sleep(0.8); st.session_state.booted = True; st.rerun()
        with st.container():
            st.markdown('<div style="background: rgba(255, 255, 255, 0.02); padding: 25px; border-radius: 8px; border: 1px solid rgba(255,255,255,0.1);">', unsafe_allow_html=True)
            u = st.text_input("Username")
            p = st.text_input("Password", type="password")
            if st.button("Access Dashboard", use_container_width=True):
                if p == MASTER_PASSWORD: 
                    st.session_state.authenticated = True; st.rerun()
                else: st.error("Access Denied")
            st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

# --- 3. CORE LOGIC (UNTOUCHED) ---
BCA_BATCHES = ["BCA 2025", "BCA 2024", "BCA 2023"]
MCA_BATCHES = ["MCA 2025", "MCA 2024"]
BLACKLIST = ["BADMINTON", "BASKETBALL", "CROSS FITNESS", "SOFT SKILL", "SWIMMING", "ZUMBA", "FREESLOT", "TABLE TENNIS", "SS ATOM", "FREE SLOT"]
ATT_COL_NAME = "Attended Hours with Approved Leave Percentage"

def apply_styles(ws, is_summary=False):
    thin = Side(style='thin', color="4D4D4D")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    crit_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid") 
    warn_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") 
    white_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font, cell.fill, cell.border = white_font, header_fill, border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 20

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if not is_summary and cell.column > 4:
                try:
                    if cell.value != "" and cell.value is not None:
                        val = float(cell.value)
                        if val < 70:
                            cell.fill = crit_fill; cell.font = white_font
                        else:
                            cell.fill = warn_fill; cell.font = Font(bold=True, color="000000")
                except: pass

def process_grid(data_df, cols, all_subjects):
    if data_df.empty: return None
    full_grid = data_df.pivot_table(index=[cols['roll'], cols['name'], cols['batch']],
                                    columns=cols['subject'], values=cols['attendance'], sort=False).reset_index()
    for sub in all_subjects:
        if sub not in full_grid.columns: full_grid[sub] = None
    sub_list = [s for s in all_subjects if s in full_grid.columns]
    theory_cols = [c for c in sub_list if not any(x in str(c).upper() for x in ["LAB", "PRACTICAL", "WORKSHOP"])]
    full_grid['Theory Avg'] = full_grid[theory_cols].mean(axis=1).round(2)
    full_grid['Final Avg'] = full_grid[sub_list].mean(axis=1).round(2)
    mask = (full_grid[sub_list] < 75).any(axis=1)
    shortage_grid = full_grid[mask].copy()
    if shortage_grid.empty: return None
    for sub in sub_list:
        shortage_grid[sub] = shortage_grid[sub].apply(lambda x: x if (pd.notnull(x) and x < 75) else "")
    shortage_grid.insert(0, 'Sl No.', range(1, len(shortage_grid) + 1))
    final_cols = ['Sl No.', cols['roll'], cols['name'], cols['batch']] + sub_list + ['Theory Avg', 'Final Avg']
    return shortage_grid[final_cols]

def run_batch_logic(df, batches, writer, cols):
    batch_summaries = []
    for group in batches:
        course, year = group.split()
        mask = (df[cols['batch']].astype(str).str.contains(course, case=False, na=False)) & \
               (df[cols['batch']].astype(str).str.contains(year, na=False))
        batch_df = df[mask].copy()
        if batch_df.empty: continue
        core_subjects = sorted([s for s in batch_df[cols['subject']].unique() if not any(b in str(s).upper() for b in BLACKLIST)])
        gen_grid = process_grid(batch_df, cols, core_subjects)
        if gen_grid is not None:
            sh_name = f"{group} ALL"[:31]
            gen_grid.to_excel(writer, sheet_name=sh_name, index=False)
            apply_styles(writer.sheets[sh_name])
            with st.expander(f"👁️ View {group} Master List"):
                st.dataframe(gen_grid, hide_index=True)
        unique_batches = sorted(batch_df[cols['batch']].unique())
        for ub in unique_batches:
            sec_df = batch_df[batch_df[cols['batch']] == ub]
            sec_grid = process_grid(sec_df, cols, core_subjects)
            if sec_grid is not None:
                sh_name = str(ub).replace("2024-2027", "").replace("2025-2028", "").strip()[:31]
                sec_grid.to_excel(writer, sheet_name=sh_name, index=False)
                apply_styles(writer.sheets[sh_name])
                batch_summaries.append({'Batch': ub, 'Count': len(sec_grid)})
    return batch_summaries

# --- 4. THE GLASS UI INTERFACE ---
st.markdown('<p class="welcome-note">Precision Analytics Dashboard</p>', unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### 🛠️ Session Control")
    if st.button("Logout of System"):
        st.session_state.authenticated = False
        st.session_state.show_logout_msg = True
        st.rerun()

uploaded_file = st.file_uploader("📂 Upload Attendance Excel", type=["xlsx"])

if uploaded_file:
    raw_head = pd.read_excel(uploaded_file, header=None).head(10)
    h_row = 0
    for i, row in raw_head.iterrows():
        if any(k in str(row.values) for k in ["Roll No", "Batch"]):
            h_row = i; break
    
    df = pd.read_excel(uploaded_file, header=h_row)
    c_map = {}
    for c in df.columns:
        cs = str(c).strip()
        if "Roll No" in cs: c_map['roll'] = c
        elif "Student Name" in cs: c_map['name'] = c
        elif "Batch" in cs: c_map['batch'] = c
        elif any(x in cs for x in ["Course", "Subject"]): c_map['subject'] = c
        elif ATT_COL_NAME in cs: c_map['attendance'] = c

    df[c_map['attendance']] = pd.to_numeric(df[c_map['attendance']], errors='coerce')
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        t1, t2, t3 = st.tabs(["💎 BCA SERIES", "🔮 MCA SERIES", "📊 COMMAND CENTER"])
        
        with t1: bca_data = run_batch_logic(df, BCA_BATCHES, writer, c_map)
        with t2: mca_data = run_batch_logic(df, MCA_BATCHES, writer, c_map)
        
        with t3:
            summary_df = pd.DataFrame(bca_data + mca_data)
            if not summary_df.empty:
                st.markdown("### 📍 System Overview")
                cols_metric = st.columns(len(summary_df))
                for idx, row in summary_df.iterrows():
                    with cols_metric[idx % len(cols_metric)]:
                        st.markdown(f"""<div class="glass-metric">
                            <div class="metric-title">{row['Batch']}</div>
                            <div class="metric-value">{row['Count']}</div>
                            <div style="color:#64748b; font-size:11px;">Shortage Detected</div>
                        </div>""", unsafe_allow_html=True)

                # --- MULTI-COLOR BAR CHART ---
                fig = px.bar(
                    summary_df, 
                    x='Batch', 
                    y='Count', 
                    color='Batch', # This enables the different colors per bar
                    title="Shortage Analytics Breakdown", 
                    template="plotly_dark",
                    color_discrete_sequence=px.colors.qualitative.Pastel # You can change this to 'Vivid', 'Prism', etc.
                )
                fig.update_layout(showlegend=False) # Hide legend to keep it clean
                st.plotly_chart(fig, use_container_width=True)

                summary_df.to_excel(writer, sheet_name='MASTER SUMMARY', index=False)
                apply_styles(writer.sheets['MASTER SUMMARY'], is_summary=True)

    st.download_button("📥 Extract Magic Report", output.getvalue(), "VMS_Analytics_Report.xlsx")

st.markdown('<div class="footer">Designed by © VMS | Secure Session Active</div>', unsafe_allow_html=True)
